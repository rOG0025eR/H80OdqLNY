# 代码生成时间: 2025-09-23 12:26:34
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

"""
Excel表格自动生成器

这个程序使用openpyxl库来创建Excel文件，并使用requests库从远程服务器获取数据。
"""

# 函数：获取远程数据
def get_remote_data(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # 检查请求是否成功
        return response.json()
    except requests.RequestException as e:
        print(f"请求失败：{e}")
        return None

# 函数：创建Excel工作簿
def create_excel_workbook():
    return Workbook()

# 函数：创建Excel工作表
def create_excel_worksheet(workbook):
    return workbook.active

# 函数：写入数据到Excel
def write_data_to_excel(ws, data):
    row_num = 1
    for row_data in data:
        column_num = 1
        for cell_data in row_data:
            ws.cell(row=row_num, column=column_num).value = cell_data
            column_num += 1
        row_num += 1

# 函数：保存Excel文件
def save_excel_file(workbook, filename):
    workbook.save(filename)

# 主程序
def main():
    # 远程数据URL
    remote_data_url = "https://api.example.com/data"
    # 要保存的Excel文件名
    excel_filename = "auto_generated_excel.xlsx"

    # 获取远程数据
    remote_data = get_remote_data(remote_data_url)
    if remote_data is None:
        print("无法获取远程数据，程序退出。")
        return

    # 创建Excel工作簿和工作表
    wb = create_excel_workbook()
    ws = create_excel_worksheet(wb)
    ws.title = "Data"

    # 写入数据到Excel
    write_data_to_excel(ws, remote_data)

    # 保存Excel文件
    save_excel_file(wb, excel_filename)

    print(f"Excel文件已保存：{excel_filename}")

if __name__ == '__main__':
    main()