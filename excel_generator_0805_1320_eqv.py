# 代码生成时间: 2025-08-05 13:20:16
import openpyxl
from openpyxl.utils import get_column_letter
from requests import get

"""
Excel表格自动生成器

这个程序使用Python和openpyxl库来创建Excel表格，并使用requests库来获取数据。
"""

class ExcelGenerator:
    def __init__(self, url, output_filename):
        """
        初始化Excel生成器
        :param url: 数据源URL
# 扩展功能模块
        :param output_filename: 输出Excel文件名
        """
        self.url = url
        self.output_filename = output_filename
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = 'Data'

    def fetch_data(self):
        """
        从指定URL获取数据
        """
        try:
            response = get(self.url)
            response.raise_for_status()  # 检查HTTP响应状态
            return response.json()
        except Exception as e:
            print(f'Error fetching data: {e}')
            return None

    def write_data_to_excel(self, data):
# 增强安全性
        """
        将数据写入Excel表格
        :param data: 要写入的数据
        """
        if not data:
            print('No data to write.')
# 扩展功能模块
            return

        # 写入标题行
        for col, key in enumerate(data[0].keys(), start=1):
            self.sheet.cell(row=1, column=col).value = key

        # 写入数据行
# 改进用户体验
        for row, item in enumerate(data, start=2):
            for col, value in enumerate(item.values(), start=1):
                self.sheet.cell(row=row, column=col).value = value

    def save_workbook(self):
        """
        保存工作簿
        """
        try:
            self.workbook.save(filename=self.output_filename)
            print(f'Excel file saved successfully: {self.output_filename}')
        except Exception as e:
            print(f'Error saving Excel file: {e}')

    def generate_excel(self):
        """
        生成Excel表格
        """
        data = self.fetch_data()
        self.write_data_to_excel(data)
        self.save_workbook()

# 使用示例
if __name__ == '__main__':
    generator = ExcelGenerator('https://api.example.com/data', 'output.xlsx')
    generator.generate_excel()